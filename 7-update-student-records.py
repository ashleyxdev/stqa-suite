from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

FILENAME = "student_records.xlsx"
SHEET_TITLE = "Students"
HEADER = ["Roll No", "Name", "Marks", "Grade"]
DATA = [
    [1,  "Amit Sharma",  72, "B"],
    [2,  "Priya Patil",  85, "A"],
    [3,  "Ravi Kumar",   60, "C"],
    [4,  "Sneha Joshi",  91, "A+"],
    [5,  "Rahul Desai",  55, "C"],
    [6,  "Pooja Mehta",  78, "B+"],
    [7,  "Arjun Nair",   88, "A"],
    [8,  "Divya Rao",    45, "D"],
    [9,  "Karan Singh",  67, "B"],
    [10, "Neha Gupta",   95, "A+"],
]

def create_sheet():
    wb = Workbook()
    ws = wb.active

    ws.title = SHEET_TITLE
    ws.append(HEADER)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFA500")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    wb.save(FILENAME)
    print("Sheet created:", FILENAME)

def insert_data():
    wb = load_workbook(FILENAME)
    ws = wb.active

    for student in DATA:
        ws.append(student)

    wb.save(FILENAME)
    print("Data inserted:", FILENAME)

def display_records():
    wb = load_workbook(FILENAME)
    ws = wb.active

    print("\nCurrent Records:")
    print("-" * 45)
    for row in ws.iter_rows(values_only=True):
        print(row)
    print("-" * 45)

def update_records():
    wb = load_workbook(FILENAME)
    ws = wb.active

    updates = {
        3: (75, "B"),
        5: (80, "A"),
        8: (70, "B"),
    }

    for roll_no, (new_marks, new_grade) in updates.items():
        row = roll_no + 1
        ws.cell(row=row, column=3).value = new_marks
        ws.cell(row=row, column=4).value = new_grade
        print(f"Updated Roll No {roll_no}: Marks={new_marks}, Grade={new_grade}")

    wb.save(FILENAME)
    print("Records updated successfully.")

# ------------------------------------

create_sheet()
insert_data()
display_records()

update_records()
display_records()