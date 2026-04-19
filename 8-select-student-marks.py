from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

FILENAME = "student_marks.xlsx"
SHEET_TITLE = "Students"
HEADER = ["Roll No", "Name", "Maths", "Physics", "Chemistry"]
DATA = [
    [1,  "Amit Sharma",  75, 55, 80],
    [2,  "Priya Patil",  50, 45, 40],
    [3,  "Ravi Kumar",   65, 70, 60],
    [4,  "Sneha Joshi",  90, 85, 88],
    [5,  "Rahul Desai",  40, 55, 58],
    [6,  "Pooja Mehta",  70, 65, 72],
    [7,  "Arjun Nair",   55, 48, 50],
    [8,  "Divya Rao",    80, 90, 75],
    [9,  "Karan Singh",  45, 62, 55],
    [10, "Neha Gupta",   95, 92, 90],
]

def create_sheet():
    wb = Workbook()
    ws = wb.active

    ws.title = SHEET_TITLE
    ws.append(HEADER)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="FFA500")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    wb.save(FILENAME)
    print("Sheet created:", FILENAME)

def insert_data():
    wb = load_workbook(FILENAME)
    ws = wb.active

    for student in DATA:
        ws.append(student)

    wb.save(FILENAME)
    print("Data inserted:", FILENAME)

def select_any():
    wb = load_workbook(FILENAME)
    ws = wb.active

    print("\nStudents who scored > 60 in AT LEAST ONE subject:")
    print("-" * 50)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        roll_no, name, maths, physics, chemistry = row
        if maths > 60 or physics > 60 or chemistry > 60:
            print(f"{roll_no}: {name} | Maths={maths}, Physics={physics}, Chemistry={chemistry}")
            count += 1

    print(f"\nTotal students = {count}")

def select_all():
    wb = load_workbook(FILENAME)
    ws = wb.active

    print("\nStudents who scored > 60 in ALL subjects:")
    print("-" * 50)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        roll_no, name, maths, physics, chemistry = row
        if maths > 60 and physics > 60 and chemistry > 60:
            print(f"{roll_no}: {name} | Maths={maths}, Physics={physics}, Chemistry={chemistry}")
            count += 1

    print(f"\nTotal students = {count}")

# ------------------------------------

create_sheet()
insert_data()
select_any()
select_all()